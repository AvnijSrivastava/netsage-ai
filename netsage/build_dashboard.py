"""
NetSage AI - Dashboard Builder
-----------------------------------------------
Merges cases.csv, ai_diagnosis.csv, human_review.csv, and
rule_checker_output.csv into a single workbook:

  Sheet 1 "RawData"  - one row per case with everything joined together
  Sheet 2 "Summary"  - counts by issue type / severity, AI-vs-human
                       agreement rate, review status breakdown, all
                       driven by formulas (COUNTIF/COUNTIFS) referencing
                       RawData so it recalculates if RawData changes
  Sheet 3 chart       - bar chart of case count by category, embedded
                       on the Summary sheet

Run:
    python build_dashboard.py
Output:
    data/dashboard.xlsx   (run scripts/recalc.py equivalent afterward)
"""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
LABEL_FONT = Font(name=FONT_NAME, bold=True)
BODY_FONT = Font(name=FONT_NAME)


def load(path):
    with open(path, newline="", encoding="utf-8") as f:
        return {r["case_id"]: r for r in csv.DictReader(f)}


def main():
    cases = load("data/cases.csv")
    ai = load("data/ai_diagnosis.csv")
    review = load("data/human_review.csv")
    rules = load("data/rule_checker_output.csv")

    case_ids = sorted(cases.keys(), key=lambda x: int(x[1:]))

    wb = Workbook()

    # ---------------- RawData sheet ----------------
    raw = wb.active
    raw.title = "RawData"
    headers = [
        "case_id", "category", "severity", "osi_layer", "concept_tag",
        "expected_fault", "ai_root_cause", "ai_confidence",
        "ai_concept_tag", "concept_tag_match",
        "rule_checker_findings", "rule_checker_hits",
        "reviewer_status", "reviewer_notes",
    ]
    raw.append(headers)
    for cell in raw[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cid in case_ids:
        c, a, r, rc = cases[cid], ai[cid], review[cid], rules[cid]
        match = "Yes" if a["concept_tag"].strip().lower() == c["concept_tag"].strip().lower() else "No"
        hits = 0 if rc["rule_findings"] == "NO_DETERMINISTIC_MATCH" else int(rc["num_findings"])
        raw.append([
            cid, c["category"], c["severity"], c["osi_layer"], c["concept_tag"],
            c["expected_fault"], a["root_cause"], a["confidence"],
            a["concept_tag"], match,
            rc["rule_findings"], hits,
            r["reviewer_status"], r["reviewer_notes"],
        ])

    widths = [8, 10, 10, 10, 22, 45, 45, 10, 22, 12, 45, 10, 14, 55]
    for i, w in enumerate(widths, start=1):
        raw.column_dimensions[get_column_letter(i)].width = w
    for row in raw.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    raw.freeze_panes = "A2"

    n = len(case_ids)
    last_row = n + 1  # header is row 1

    # ---------------- Summary sheet ----------------
    ws = wb.create_sheet("Summary")
    ws["A1"] = "NetSage AI - Case & Review Dashboard"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    # Category counts
    ws["A3"] = "Cases by Issue Type"
    ws["A3"].font = LABEL_FONT
    categories = ["VLAN", "Gateway", "DHCP", "DNS", "Routing", "ACL", "NAT", "Wireless"]
    ws["A4"], ws["B4"] = "Category", "Count"
    ws["A4"].font = HEADER_FONT
    ws["B4"].font = HEADER_FONT
    ws["A4"].fill = HEADER_FILL
    ws["B4"].fill = HEADER_FILL
    row = 5
    cat_start_row = row
    for cat in categories:
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=f'=COUNTIF(RawData!B2:B{last_row},A{row})')
        row += 1
    cat_end_row = row - 1

    # Severity counts
    sev_col_start = 4  # column D
    ws.cell(row=4, column=sev_col_start, value="Severity").font = HEADER_FONT
    ws.cell(row=4, column=sev_col_start).fill = HEADER_FILL
    ws.cell(row=4, column=sev_col_start + 1, value="Count").font = HEADER_FONT
    ws.cell(row=4, column=sev_col_start + 1).fill = HEADER_FILL
    severities = ["High", "Medium", "Low"]
    srow = 5
    for sev in severities:
        ws.cell(row=srow, column=sev_col_start, value=sev)
        ws.cell(row=srow, column=sev_col_start + 1,
                value=f'=COUNTIF(RawData!C2:C{last_row},{get_column_letter(sev_col_start)}{srow})')
        srow += 1

    # Review status breakdown
    ws.cell(row=10, column=1, value="Human Review Outcomes").font = LABEL_FONT
    ws.cell(row=11, column=1, value="Status").font = HEADER_FONT
    ws.cell(row=11, column=1).fill = HEADER_FILL
    ws.cell(row=11, column=2, value="Count").font = HEADER_FONT
    ws.cell(row=11, column=2).fill = HEADER_FILL
    statuses = ["Accepted", "Edited", "Rejected"]
    row = 12
    status_start = row
    for st in statuses:
        ws.cell(row=row, column=1, value=st)
        ws.cell(row=row, column=2, value=f'=COUNTIF(RawData!M2:M{last_row},A{row})')
        row += 1
    status_end = row - 1

    ws.cell(row=row + 1, column=1, value="Total cases").font = LABEL_FONT
    ws.cell(row=row + 1, column=2, value=f'=COUNTA(RawData!A2:A{last_row})')

    ws.cell(row=row + 2, column=1, value="AI vs Human Agreement Rate").font = LABEL_FONT
    ws.cell(row=row + 2, column=2,
            value=f'=COUNTIF(RawData!J2:J{last_row},"Yes")/COUNTA(RawData!A2:A{last_row})')
    ws.cell(row=row + 2, column=2).number_format = "0.0%"

    ws.cell(row=row + 3, column=1, value="Accepted-Without-Edit Rate").font = LABEL_FONT
    ws.cell(row=row + 3, column=2,
            value=f'=COUNTIF(RawData!M2:M{last_row},"Accepted")/COUNTA(RawData!A2:A{last_row})')
    ws.cell(row=row + 3, column=2).number_format = "0.0%"

    ws.cell(row=row + 4, column=1, value="Cases with Deterministic Rule Hit").font = LABEL_FONT
    ws.cell(row=row + 4, column=2,
            value=f'=COUNTIF(RawData!L2:L{last_row},">0")')

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10

    # Chart: cases by category
    chart = BarChart()
    chart.title = "Cases by Issue Type"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Category"
    data = Reference(ws, min_col=2, min_row=4, max_row=cat_end_row)
    cats = Reference(ws, min_col=1, min_row=cat_start_row, max_row=cat_end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 16
    chart.height = 9
    ws.add_chart(chart, "G4")

    # Chart: review outcomes
    chart2 = BarChart()
    chart2.title = "Human Review Outcomes"
    chart2.y_axis.title = "Count"
    data2 = Reference(ws, min_col=2, min_row=11, max_row=status_end)
    cats2 = Reference(ws, min_col=1, min_row=status_start, max_row=status_end)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.width = 16
    chart2.height = 9
    ws.add_chart(chart2, "G21")

    os.makedirs("data", exist_ok=True)
    out_path = "data/dashboard.xlsx"
    wb.save(out_path)
    print(f"Wrote dashboard to {out_path}")


if __name__ == "__main__":
    main()
